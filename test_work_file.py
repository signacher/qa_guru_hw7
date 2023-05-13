import csv
import zipfile
from openpyxl import load_workbook
from selenium import webdriver
from selene import browser
import os.path
import requests
from pypdf import PdfReader
import xlrd
from os_path_scripts import RESOURCES_PATH, TMP_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp
def test_download_file_from_browser():
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": TMP_PATH,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()


# TODO сохранять и читать из tmp, использовать универсальный путь
def test_downloaded_file_size():
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    logo_file_path = os.path.join(TMP_PATH, 'selenium_logo.png')
    r = requests.get(url)

    with open(logo_file_path, 'wb') as file:
        file.write(r.content)

    size = os.path.getsize(logo_file_path)

    assert size == 30803


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_csv_file():
    users = []
    csv_file_path=os.path.join(RESOURCES_PATH,'eggs.csv')
    with open(csv_file_path, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(csv_file_path) as csvfile:
        csvreader = csv.reader(csvfile)

        for row in csvreader:
            users.append(row)
            print(row)
    assert users[0] == ['Anna', 'Pavel', 'Peter']


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_pdf_file():
    pdf_file_path = os.path.join(RESOURCES_PATH,'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(pdf_file_path)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()

    print(page)
    print(number_of_pages)
    print(text)

    assert number_of_pages == 412


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xls_file():
    xls_file_path = os.path.join(RESOURCES_PATH,'file_example_XLS_10.xls')
    book = xlrd.open_workbook(xls_file_path)

    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))

    assert book.nsheets == 1
    assert sheet.cell_value(rowx=9, colx=7) == 6548.0


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx_file():
    xlsx_file_path = os.path.join(RESOURCES_PATH,'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active

    print(sheet.cell(row=3, column=2).value)

    assert (sheet.cell(row=4, column=5).value) == 'France'

def test_zip_file():
    zip_file_path = os.path.join(RESOURCES_PATH,'hello.zip')
    with zipfile.ZipFile(zip_file_path) as hello_zip:
        print(hello_zip.namelist())
        text = hello_zip.read('Hello.txt')
        print(text)

    assert hello_zip.namelist() == ['Hello.txt']
    assert text ==b'Hello world\n'


def test_extract_zip_file():
    zip_file_path = os.path.join(RESOURCES_PATH, 'hello.zip')
    with zipfile.ZipFile(zip_file_path) as hello_zip:
        hello_zip.extractall(TMP_PATH)
        tmp_name_list = os.listdir(TMP_PATH)
        print(tmp_name_list)
        tmp_extract_file_path = os.path.join(TMP_PATH, 'Hello.txt')

    assert os.path.exists(tmp_extract_file_path) == True


def test_add_to_zip():
    zip_file = 'zip_test_file.zip'

    with zipfile.ZipFile(zip_file, 'w') as zip_files:
         for file in os.listdir(RESOURCES_PATH):
             file_path = os.path.join(RESOURCES_PATH, file)
             zip_files.write(file_path, file)

    with zipfile.ZipFile(zip_file, "r") as zip_files:
        for file in os.listdir(RESOURCES_PATH):
            assert file in zip_files.namelist()

    os.remove(zip_file)
